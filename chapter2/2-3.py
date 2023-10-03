import openpyxl as px

wb = px.Workbook()

ws = wb["Sheet"]
ws.title = "しーと"
print(wb.sheetnames)

ws["A1"] = "ABCDE"
ws.cell(row=2, column=1, value="12345")
ws.cell(row=1, column=1).fill = px.styles.PatternFill(patternType="solid", fgColor="aaffff")

for i in range(1, 10):
    ws.cell(row=5, column=1+i, value=i).fill = px.styles.PatternFill(patternType="solid", fgColor="aaaaaa")
    # 5行目から右に1から9を並べ、セルの色を濃いグレーにする。
    ws.cell(row=5+i, column=1, value=i).fill = px.styles.PatternFill(patternType="solid", fgColor="aaaaaa")
    # 5行目から下に1から9を並べ、セルの色を濃いグレーにする。

for x in range(1, 10):
    for y in range(1, 10):
        ws.cell(row=5+x, column=1+y, value=x*y).fill = px.styles.PatternFill(patternType="solid", fgColor="cccccc")
        # 各セルに掛け算の答えを順番に入れていく。
        
wb.save("Book1.csv")